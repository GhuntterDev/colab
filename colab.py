# app.py — Avaliação de Colaboradores (CSV + Google Sheets)
# Colunas por posição:
# A=Data, B=Setor, C=Colaborador, D=Velocidade, F=Atendimento, H=Qualidade, J=Ajuda, M=Avaliador

import io
import re
from datetime import datetime
from typing import List, Tuple

import numpy as np
import pandas as pd
import streamlit as st
from gspread.exceptions import APIError

# ---------------------------------
# CONFIG & DEBUG
# ---------------------------------
st.set_page_config(page_title="Avaliação de Colaboradores", layout="wide")
DEBUG_LOGS = False  # Desativado - sistema funcionando
def _log(msg):
    if DEBUG_LOGS:
        st.write(f"🔍 DEBUG: {msg}")
        # Remove todos os emojis Unicode para o console do Windows
        import re
        console_msg = re.sub(r'[^\x00-\x7F]+', '[EMOJI]', msg)
        print(f"DEBUG: {console_msg}")  # Console sem emojis

# Adicionado mapeamento de lojas por região
MAPEAMENTO_REGIOES = {
    # Rio de Janeiro
    "Carioca": "RJ",
    "Santa Cruz": "RJ",
    "Mesquita": "RJ",
    "Nilópolis": "RJ",
    "Madureira": "RJ",
    "Bonsucesso": "RJ",
    # São Paulo
    "Taboão": "SP",
    "São Bernardo": "SP",
    "Santo André": "SP",
    "Mauá": "SP",
    "MDC São Mateus": "SP",
    "CDM São Mateus": "SP",
}

def get_regiao(loja: str) -> str:
    """Retorna a região da loja ou 'Outra' se não mapeada"""
    return MAPEAMENTO_REGIOES.get(loja, "Outra")

# Mapeamento por letra
NOTAS_COLS = {"Velocidade": "D", "Atendimento": "F", "Qualidade": "H", "Ajuda": "J"}
COL_DATA, COL_SETOR, COL_NOME, COL_AVALIADOR = "A", "B", "C", "M"


def normalize_colnames(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [c.encode("utf-8").decode("utf-8-sig").strip() if isinstance(c, str) else c for c in df.columns]
    return df

def get_col_by_letter(df: pd.DataFrame, letter: str) -> str:
    idx = ord(letter.upper()) - ord("A")
    if not (0 <= idx < len(df.columns)):
        raise ValueError(f"Letra de coluna inválida: {letter}")
    return df.columns[idx]

def parse_datetime_ptbr(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce", dayfirst=True)

def cast_notas_safe(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    for c in cols:
        if c not in df.columns:
            df[c] = 0.0
            continue
        obj = df[c]
        if isinstance(obj, pd.DataFrame):
            obj = obj.iloc[:, 0]
        df[c] = pd.to_numeric(obj, errors="coerce").fillna(0.0)
    return df

def infer_loja_from_filename(name: str) -> str:
    if not name: return "Desconhecida"
    m = re.search(r"-\s*([^-()]+?)\s*(?:\(|-|\.|$)", name)
    if m: return m.group(1).strip()
    tokens = re.findall(r"[A-Za-zÀ-ÖØ-öø-ÿ]+", name)
    return tokens[-1] if tokens else "Desconhecida"

def build_summary(df: pd.DataFrame, by_cols: List[str]) -> pd.DataFrame:
    notas = ["Velocidade", "Atendimento", "Qualidade", "Ajuda"]
    work = df.copy()
    work["Avaliações"] = 1
    if by_cols:
        means = work.groupby(by_cols, dropna=False)[notas].mean(numeric_only=True)
        counts = work.groupby(by_cols, dropna=False)["Avaliações"].sum()
        out = means.join(counts).reset_index()
    else:
        out = pd.DataFrame({
            "Velocidade": [work["Velocidade"].mean()],
            "Atendimento": [work["Atendimento"].mean()],
            "Qualidade": [work["Qualidade"].mean()],
            "Ajuda": [work["Ajuda"].mean()],
            "Avaliações": [work["Avaliações"].sum()],
        })
    out["Média Geral"] = out[["Velocidade","Atendimento","Qualidade","Ajuda"]].mean(axis=1)
    out[["Velocidade","Atendimento","Qualidade","Ajuda","Média Geral"]] = \
        out[["Velocidade","Atendimento","Qualidade","Ajuda","Média Geral"]].round(2)
    if by_cols:
        out = out.sort_values("Média Geral", ascending=False)
    return out


REQUIRED_GCP_FIELDS = {
    "type","project_id","private_key_id","private_key",
    "client_email","client_id","auth_uri","token_uri",
    "auth_provider_x509_cert_url","client_x509_cert_url"
}

def _normalize_sheet_id(maybe_url_or_id: str) -> str:
    s = (maybe_url_or_id or "").strip()
    m = re.search(r"/d/([a-zA-Z0-9\-_]+)", s)
    return m.group(1) if m else s

def _load_sa_creds_from_secrets() -> dict:
    _log("🔑 Tentando carregar credenciais...")
    
    # Primeiro, tentar carregar do arquivo JSON
    try:
        import json
        import os
        json_path = "service_account.json"
        if os.path.exists(json_path):
            _log("📄 Carregando credenciais do arquivo JSON...")
            with open(json_path, 'r') as f:
                creds = json.load(f)
            _log(f"✅ Arquivo JSON carregado com sucesso")
            _log(f"📧 Service Account: {creds.get('client_email', 'NÃO ENCONTRADO')}")
            return creds
    except Exception as e:
        _log(f"❌ Erro ao carregar arquivo JSON: {e}")
    
    # Se não conseguir carregar do JSON, tentar do secrets.toml
    _log("📄 Tentando carregar credenciais do secrets.toml...")
    try:
        creds = dict(st.secrets["gcp_service_account"])
        _log(f"✅ Seção [gcp_service_account] encontrada")
        _log(f"📧 Service Account: {creds.get('client_email', 'NÃO ENCONTRADO')}")
        
        # corrige chave salva com \n literal (duas barras) -> quebra real
        pk = creds.get("private_key", "")
        if "\\n" in pk and "\n" not in pk:
            _log("🔧 Corrigindo formatação da chave privada...")
            creds["private_key"] = pk.replace("\\n", "\n")
        
        missing = [k for k in REQUIRED_GCP_FIELDS if k not in creds or not str(creds[k]).strip()]
        if missing:
            present = sorted(creds.keys())
            _log(f"❌ Credenciais incompletas. Faltando: {missing}")
            _log(f"📋 Campos presentes: {present}")
            raise RuntimeError(f"Credenciais incompletas. Faltando: {missing}. Presentes: {present}.")
        
        _log("✅ Credenciais do secrets.toml validadas com sucesso")
        return creds
        
    except Exception as e:
        _log(f"❌ Erro ao carregar secrets: {e}")
        raise RuntimeError("Não foi possível carregar credenciais nem do arquivo JSON nem do secrets.toml")

def _open_sheet_by_id(maybe_url_or_id: str):
    import gspread
    _log(f"🔗 URL/ID recebido: {maybe_url_or_id}")
    sid = _normalize_sheet_id(maybe_url_or_id)
    _log(f"🔎 ID normalizado: `{sid}`")
    
    creds = _load_sa_creds_from_secrets()
    _log(f"📧 Service account: {creds.get('client_email')}")
    
    _log("🔌 Conectando ao Google Sheets API...")
    try:
        gc = gspread.service_account_from_dict(creds)
        _log("✅ Conectado ao Google Sheets API")
    except Exception as e:
        _log(f"❌ Erro ao conectar com Google Sheets API: {e}")
        raise
    
    _log(f"📋 Tentando abrir planilha com ID: {sid}")
    try:
        sh = gc.open_by_key(sid)
        _log(f"✅ Planilha aberta: {sh.title}")
        return sh
    except APIError as e:
        msg = str(e)
        _log(f"❌ APIError ao abrir planilha: {msg}")
        if "404" in msg:
            error_msg = "Planilha não encontrada (404). Verifique ID (trecho entre /d/ e /edit) e se a planilha foi compartilhada com a service account (Leitor)."
            _log(f"❌ {error_msg}")
            raise RuntimeError(error_msg)
        if "403" in msg:
            error_msg = "Sem permissão (403). Compartilhe a planilha com a service account como Leitor."
            _log(f"❌ {error_msg}")
            raise RuntimeError(error_msg)
        _log(f"❌ Erro não tratado: {msg}")
        raise
    except Exception as e:
        _log(f"❌ Erro inesperado ao abrir planilha: {e}")
        _log(f"❌ Tipo do erro: {type(e).__name__}")
        raise

@st.cache_data(ttl=300)  # Cache por 5 minutos
def _fetch_from_gsheets(spreadsheet_id: str) -> List[Tuple[str, pd.DataFrame]]:
    import time
    try:
        sh = _open_sheet_by_id(spreadsheet_id)
        ws_list = sh.worksheets()
        _log(f"📑 {sh.title}: {len(ws_list)} aba(s) detectadas")
        dfs: List[Tuple[str, pd.DataFrame]] = []
        
        for i, ws in enumerate(ws_list):
            _log(f"➡️ Lendo aba: {ws.title}")
            
            # Pequena pausa entre requisições para evitar quota
            if i > 0:
                time.sleep(0.5)
            
            values = ws.get_all_values()
            if not values:
                _log("   (vazia)")
                continue
            if len(values) >= 2:
                df = pd.DataFrame(values[1:], columns=values[0])
            else:
                df = pd.DataFrame(values)
            dfs.append((ws.title, df))
        return dfs
    except Exception as e:
        if "429" in str(e) or "quota" in str(e).lower():
            st.warning("⚠️ Limite de requisições da API excedido. Aguarde alguns minutos e tente novamente.")
            st.info("💡 Dica: O cache foi ativado para reduzir requisições.")
        raise


st.sidebar.header("Configurações")
# Link padrão do Google Sheets
SPREADSHEET_URL = "https://docs.google.com/spreadsheets/d/196mkyj8XPLouscoiJqmarCJ4H5N7ANnAhmH6V-uXSlw/edit?gid=2014063945#gid=2014063945"
frames: List[pd.DataFrame] = []

# Sempre usar Google Sheets com o link padrão
spreadsheet_in = SPREADSHEET_URL
_log(f"🚀 Iniciando carregamento de dados...")
_log(f"🔗 URL da planilha: {SPREADSHEET_URL}")
col1, col2 = st.sidebar.columns(2)
with col1:
    if st.button("↻ Atualizar", help="Atualiza os dados da planilha"):
        st.cache_data.clear()
        st.rerun()
with col2:
    if st.button("⏰ Cache", help="Limpa o cache (use se dados não atualizaram)"):
        st.cache_data.clear()
        st.success("Cache limpo!")
if spreadsheet_in:
        try:
            _log("📊 Tentando buscar dados do Google Sheets...")
            sheets = _fetch_from_gsheets(spreadsheet_in)
            for title, df in sheets:
                df = normalize_colnames(df)
                # mapear por posição
                try:
                    col_data = get_col_by_letter(df, COL_DATA)
                    col_setor = get_col_by_letter(df, COL_SETOR)
                    col_nome  = get_col_by_letter(df, COL_NOME)
                    col_vel = get_col_by_letter(df, NOTAS_COLS["Velocidade"])
                    col_atd = get_col_by_letter(df, NOTAS_COLS["Atendimento"])
                    col_qlt = get_col_by_letter(df, NOTAS_COLS["Qualidade"])
                    col_ajd = get_col_by_letter(df, NOTAS_COLS["Ajuda"])
                    col_avaliador = get_col_by_letter(df, COL_AVALIADOR)  # M
                except Exception as e:
                    st.warning(f"Aba '{title}': problema ao mapear colunas — {e}")
                    continue

                rec = pd.DataFrame({
                    "Data": df[col_data],
                    "Setor": df[col_setor],
                    "Colaborador": df[col_nome],
                    "Velocidade": df[col_vel],
                    "Atendimento": df[col_atd],
                    "Qualidade": df[col_qlt],
                    "Ajuda": df[col_ajd],
                    "Avaliador": df[col_avaliador],  # novo
                })
                rec["Data"] = parse_datetime_ptbr(rec["Data"])
                rec = cast_notas_safe(rec, ["Velocidade","Atendimento","Qualidade","Ajuda"])

                # ignora linhas sem nome de colaborador
                rec = rec[rec["Colaborador"].notna()].copy()
                rec["Colaborador"] = rec["Colaborador"].astype(str).str.strip()
                rec = rec[(rec["Colaborador"] != "") & 
                          (rec["Colaborador"].str.lower() != "nan") & 
                          (rec["Colaborador"].str.lower() != "none")]

                # limpando avaliador (mantenho vazios? se não quiser, descomente 2 linhas abaixo)
                rec["Avaliador"] = rec["Avaliador"].astype(str).str.strip()
                # rec = rec[~rec["Avaliador"].str.lower().isin(["", "nan", "none"])]

                # novas colunas de data/hora
                rec["Data_dia"] = rec["Data"].dt.date
                rec["Hora"] = rec["Data"].dt.strftime("%H:%M")
                rec["Hora_num"] = rec["Data"].dt.hour

                rec["Loja"] = title
                # Adicionando coluna de Região baseada na loja
                rec["Região"] = rec["Loja"].apply(get_regiao)
                frames.append(rec)
        except Exception as e:
            _log(f"❌ ERRO CAPTURADO: {e}")
            _log(f"❌ Tipo do erro: {type(e).__name__}")
            st.error(f"Erro ao carregar Google Sheets: {e}")
            st.error(f"Tipo do erro: {type(e).__name__}")
            # Mostrar mais detalhes do erro
            import traceback
            _log(f"❌ Traceback completo: {traceback.format_exc()}")


if not frames:
    st.title("📊 Avaliação de Colaboradores")
    st.info("Carregando dados do Google Sheets...")
    st.stop()

data = pd.concat(frames, ignore_index=True)

# limites de data
if data["Data"].notna().any():
    date_min = pd.to_datetime(data["Data"].min()).date()
    date_max = pd.to_datetime(data["Data"].max()).date()
else:
    today = datetime.today().date()
    date_min = date_max = today

st.sidebar.subheader("Filtros")
# Corrigir problema quando date_min == date_max
if date_min == date_max:
    # Se há apenas uma data, usar apenas essa data
    d_ini = st.sidebar.date_input("Data", value=date_min, min_value=date_min, max_value=date_max)
    d_fim = d_ini
else:
    # Se há período, usar range
    d_ini, d_fim = st.sidebar.date_input("Período", value=(date_min, date_max), min_value=date_min, max_value=date_max)
st.sidebar.caption("💡 Dica: Selecione uma data para ver apenas esse dia, ou duas datas para ver o período.")

regiao_sel = st.sidebar.selectbox("Filtrar por região", options=["Todos", "SP", "RJ"], index=0)

# Selectbox para ver loja específica
loja_especifica = st.sidebar.selectbox(
    "Ver loja específica",
    options=["Todas as lojas"] + sorted([x for x in data["Loja"].dropna().unique()]),
    index=0,
    help="Selecione uma loja para visualizar apenas ela, ou 'Todas as lojas' para usar o filtro múltiplo abaixo"
)

# Multiselect só é usado se "Todas as lojas" estiver selecionado
if loja_especifica == "Todas as lojas":
    lojas_sel = st.sidebar.multiselect("Filtrar por loja",
                                       options=sorted([x for x in data["Loja"].dropna().unique()]),
                                       default=sorted([x for x in data["Loja"].dropna().unique()]))
else:
    # Se uma loja específica foi selecionada, usar apenas ela
    lojas_sel = [loja_especifica]
    st.sidebar.info(f"Visualizando apenas: {loja_especifica}")

setores_sel = st.sidebar.multiselect("Filtrar por setor",
                                     options=sorted([x for x in data["Setor"].dropna().unique()]),
                                     default=sorted([x for x in data["Setor"].dropna().unique()]))

# filtro de hora
hora_ini, hora_fim = st.sidebar.slider("Faixa de hora do dia", min_value=0, max_value=23, value=(0, 23), step=1)

mask = pd.Series(True, index=data.index)
# Filtro de data - aceita uma data ou período
if d_ini and d_fim:
    if d_ini == d_fim:
        # Se apenas uma data foi selecionada, filtrar apenas essa data
        mask &= data["Data"].dt.date == d_ini
    else:
        # Se período foi selecionado, filtrar o período
        mask &= data["Data"].between(pd.to_datetime(d_ini),
                                     pd.to_datetime(d_fim) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1),
                                     inclusive="both")
elif d_ini:
    # Se apenas data inicial foi selecionada
    mask &= data["Data"].dt.date == d_ini
if regiao_sel != "Todos":
    mask &= data["Região"] == regiao_sel
if lojas_sel:
    mask &= data["Loja"].isin(lojas_sel)
if setores_sel:
    mask &= data["Setor"].isin(setores_sel)
if "Hora_num" in data.columns:
    mask &= data["Hora_num"].between(hora_ini, hora_fim)

df_f = data.loc[mask].copy()

st.title("📊 Avaliação de Colaboradores")

# KPIs
col1, col2, col3, col4, col5 = st.columns(5)
with col1: st.metric("Avaliações (linhas)", int(df_f.shape[0]))
with col2: st.metric("Média Velocidade", round(df_f["Velocidade"].mean() if len(df_f) else 0, 2))
with col3: st.metric("Média Atendimento", round(df_f["Atendimento"].mean() if len(df_f) else 0, 2))
with col4: st.metric("Média Qualidade", round(df_f["Qualidade"].mean() if len(df_f) else 0, 2))
with col5: st.metric("Média Ajuda", round(df_f["Ajuda"].mean() if len(df_f) else 0, 2))

# Prévia
# Adicionando coluna Região na prévia
cols_preview = [c for c in ["Data_dia", "Hora", "Região", "Loja", "Setor", "Colaborador", "Avaliador",
                            "Velocidade", "Atendimento", "Qualidade", "Ajuda"] if c in df_f.columns]
st.markdown("### 🔎 Prévia dos dados filtrados")
st.dataframe(df_f.sort_values(["Data_dia","Hora"], ascending=False)[cols_preview], use_container_width=True)

st.markdown("---")
st.subheader("👤 Resumo por Pessoa")
# Incluindo Região no resumo por pessoa
resumo_pessoa = build_summary(df_f, by_cols=["Colaborador","Região","Loja","Setor"])
st.dataframe(resumo_pessoa, use_container_width=True)

st.markdown("### 🏬 Resumo por Setor")
# Incluindo Região no resumo por setor
resumo_setor = build_summary(df_f, by_cols=["Setor","Região","Loja"])
st.dataframe(resumo_setor, use_container_width=True)

st.markdown("### 🏪 Resumo por Loja")
# Incluindo Região no resumo por loja
resumo_loja = build_summary(df_f, by_cols=["Região","Loja"])
st.dataframe(resumo_loja, use_container_width=True)

# Adicionado novo resumo por região
st.markdown("### 🗺️ Resumo por Região")
resumo_regiao = build_summary(df_f, by_cols=["Região"])
st.dataframe(resumo_regiao, use_container_width=True)

st.markdown("### 🌐 Resumo Geral (todas as lojas)")
resumo_geral = build_summary(df_f, by_cols=[])
st.dataframe(resumo_geral, use_container_width=True)


st.markdown("### 🧑‍⚖️ Ranking de avaliadores (quem mais faz avaliações)")
if "Avaliador" in df_f.columns and not df_f.empty:
    # Incluindo Região no ranking de avaliadores
    rank_av = (df_f.groupby(["Avaliador","Região","Loja"])
                    .size()
                    .reset_index(name="Avaliações feitas")
                    .sort_values(["Avaliações feitas","Avaliador"], ascending=[False, True]))
    st.dataframe(rank_av, use_container_width=True)

    pivot_av = (rank_av.pivot(index="Avaliador", columns="Loja", values="Avaliações feitas")
                        .fillna(0).astype(int))
    st.markdown("#### 🔄 Distribuição por loja (avaliadores x lojas)")
    st.dataframe(pivot_av, use_container_width=True)

    try:
        import altair as alt
        chart_av = alt.Chart(rank_av).mark_bar().encode(
            x=alt.X("Avaliador:N", sort="-y"),
            y=alt.Y("Avaliações feitas:Q"),
            color="Loja:N",
            tooltip=["Avaliador","Região","Loja","Avaliações feitas"]
        ).properties(height=380)
        st.altair_chart(chart_av, use_container_width=True)
    except Exception:
        pass

    st.download_button(
        "Baixar CSV (ranking de avaliadores)",
        data=rank_av.to_csv(index=False).encode("utf-8-sig"),
        file_name="ranking_avaliadores.csv",
        mime="text/csv",
    )
else:
    st.info("Sem dados de 'Avaliador' para este filtro.")


st.markdown("### ⏰ Volume por hora (por loja)")
if not df_f.empty and "Hora_num" in df_f.columns:
    # Incluindo Região no volume por hora
    por_hora = df_f.groupby(["Região","Loja","Hora_num"]).size().reset_index(name="Avaliações")
    pivot = por_hora.pivot(index="Hora_num", columns="Loja", values="Avaliações").fillna(0).astype(int).sort_index()
    st.dataframe(pivot, use_container_width=True)

    try:
        import altair as alt
        chart = alt.Chart(por_hora).mark_bar().encode(
            x=alt.X("Hora_num:O", title="Hora do dia (0–23)"),
            y=alt.Y("Avaliações:Q"),
            color="Loja:N",
            tooltip=["Região","Loja","Hora_num","Avaliações"]
        ).properties(height=360)
        st.altair_chart(chart, use_container_width=True)
    except Exception:
        pass

    st.download_button(
        "Baixar CSV (volume por hora x loja)",
        data=por_hora.to_csv(index=False).encode("utf-8-sig"),
        file_name="volume_por_hora_por_loja.csv",
        mime="text/csv",
    )
else:
    st.info("Não há dados filtrados para montar o relatório por hora.")


def gerar_relatorio_excel_por_loja(df_in: pd.DataFrame) -> bytes:
    import io
    # Incluindo Região no relatório Excel
    base = build_summary(df_in, by_cols=["Colaborador","Região","Loja","Setor"]).copy()
    base = base.sort_values("Média Geral", ascending=False)
    cols = ["Colaborador","Região","Loja","Setor","Velocidade","Atendimento","Qualidade","Ajuda","Avaliações","Média Geral"]
    base = base[[c for c in cols if c in base.columns]]

    # Ranking de avaliadores (aba extra)
    rank_av = (df_in.groupby(["Avaliador","Região","Loja"])
                    .size()
                    .reset_index(name="Avaliações feitas")
                    .sort_values(["Avaliações feitas","Avaliador"], ascending=[False, True]))

    buf = io.BytesIO()
    try:
        import xlsxwriter  # noqa
        engine = "xlsxwriter"
    except Exception:
        engine = "openpyxl"

    with pd.ExcelWriter(buf, engine=engine) as writer:
        if engine == "xlsxwriter":
            workbook  = writer.book
            header_fmt = workbook.add_format({
                "bold": True, "bg_color": "#B7B7B7", "font_color": "#000000",
                "align": "center", "valign": "vcenter", "border": 1,
            })
            num2_fmt   = workbook.add_format({"num_format": "0.00", "border": 1, "align": "center", "valign": "vcenter"})
            int0_fmt   = workbook.add_format({"num_format": "0",    "border": 1, "align": "center", "valign": "vcenter"})
            text_fmt   = workbook.add_format({"border": 1, "align": "center", "valign": "vcenter"})

            # abas por loja
            for loja, g in base.groupby("Loja", dropna=False):
                sheet = str(loja) if pd.notna(loja) else "(Sem Loja)"
                g = g.copy().sort_values("Média Geral", ascending=False)
                g.to_excel(writer, index=False, sheet_name=sheet)
                ws = writer.sheets[sheet]

                widths = {
                    "Colaborador": 36, "Região": 10, "Loja": 18, "Setor": 20,
                    "Velocidade": 12, "Atendimento": 12, "Qualidade": 12, "Ajuda": 10,
                    "Avaliações": 12, "Média Geral": 12,
                }
                for c_idx, col_name in enumerate(g.columns):
                    ws.set_column(c_idx, c_idx, widths.get(col_name, 14), text_fmt)
                    ws.write(0, c_idx, col_name, header_fmt)

                col_map = {name: idx for idx, name in enumerate(g.columns)}
                for nm in ["Velocidade","Atendimento","Qualidade","Ajuda","Média Geral"]:
                    if nm in col_map:
                        ws.set_column(col_map[nm], col_map[nm], widths.get(nm, 12), num2_fmt)
                if "Avaliações" in col_map:
                    ws.set_column(col_map["Avaliações"], col_map["Avaliações"], widths.get("Avaliações", 10), int0_fmt)

                ws.autofilter(0, 0, len(g), len(g.columns)-1)
                ws.freeze_panes(1, 0)
                for r in range(1, len(g)+1): ws.set_row(r, 18)

                if "Média Geral" in col_map:
                    c = col_map["Média Geral"]
                    ws.conditional_format(1, c, len(g), c, {
                        "type": "3_color_scale",
                        "min_color": "#FCA5A5", "mid_color": "#FDE68A", "max_color": "#86EFAC",
                    })

            # aba extra: avaliadores
            sheet = "Avaliadores"
            rank_av.to_excel(writer, index=False, sheet_name=sheet)
            ws = writer.sheets[sheet]
            for c_idx, col_name in enumerate(rank_av.columns):
                ws.set_column(c_idx, c_idx, 24 if col_name in ["Avaliador","Loja"] else 18, text_fmt)
                ws.write(0, c_idx, col_name, header_fmt)
            if "Avaliações feitas" in rank_av.columns:
                ci = list(rank_av.columns).index("Avaliações feitas")
                ws.set_column(ci, ci, 18, int0_fmt)
            ws.autofilter(0, 0, len(rank_av), max(0, len(rank_av.columns)-1))
            ws.freeze_panes(1, 0)

        else:
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.formatting.rule import ColorScaleRule

            # abas por loja
            for loja, g in base.groupby("Loja", dropna=False):
                sheet = str(loja) if pd.notna(loja) else "(Sem Loja)"
                g = g.copy().sort_values("Média Geral", ascending=False)
                g.to_excel(writer, index=False, sheet_name=sheet)
                ws = writer.sheets[sheet]

                header_fill = PatternFill("solid", fgColor="B7B7B7")
                header_font = Font(bold=True, color="000000")
                thin = Side(style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                widths = {
                    "Colaborador": 36, "Região": 10, "Loja": 18, "Setor": 20,
                    "Velocidade": 12, "Atendimento": 12, "Qualidade": 12, "Ajuda": 10,
                    "Avaliações": 12, "Média Geral": 12,
                }
                for i, col in enumerate(g.columns, start=1):
                    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = widths.get(col, 14)

                ws.auto_filter.ref = ws.dimensions
                ws.freeze_panes = "A2"

                try:
                    mg_col = list(g.columns).index("Média Geral") + 1
                    ws.conditional_formatting.add(
                        f"{ws.cell(row=2, column=mg_col).coordinate}:{ws.cell(row=ws.max_row, column=mg_col).coordinate}",
                        ColorScaleRule(start_type='min', start_color='FCA5A5',
                                       mid_type='percentile', mid_value=50, mid_color='FDE68A',
                                       end_type='max', end_color='86EFAC')
                    )
                except ValueError:
                    pass

            # aba extra: avaliadores
            sheet = "Avaliadores"
            rank_av.to_excel(writer, index=False, sheet_name=sheet)
            ws = writer.sheets[sheet]
            header_fill = PatternFill("solid", fgColor="B7B7B7")
            header_font = Font(bold=True, color="000000")
            thin = Side(style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            for i, col in enumerate(rank_av.columns, start=1):
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 24 if col in ["Avaliador","Loja"] else 18
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

    buf.seek(0)
    return buf.getvalue()

st.markdown("---")
st.subheader("📄 Exportar relatório por loja (Excel)")
if st.button("⬇️ Gerar relatório Excel (uma aba por loja + Avaliadores)"):
    xlsx_bytes = gerar_relatorio_excel_por_loja(df_f)
    st.download_button(
        label="Baixar relatório.xlsx",
        data=xlsx_bytes,
        file_name="relatorio_por_loja.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

with st.expander("ℹ️ Dicas e validações"):
    st.markdown(
        """
- **Vazios contam como 0** nas colunas de nota.
- **Colaborador vazio** é ignorado.
- A coluna **Data** é interpretada como **dd/mm/aaaa** (com ou sem hora).
- Em **Google Sheets**, cada **aba** vira uma **Loja**.
- Filtro de **hora** permite ver horários mais ativos por loja.
- Coluna **M = Avaliador** permite acompanhar quem mais avalia.
- **Filtro de Região**: RJ (Carioca, Santa Cruz, Mesquita, Nilópolis, Madureira, Bonsucesso) e SP (Taboão, São Bernardo, Santo André, Mauá, MDC São Mateus, CDM São Mateus).
        """
    )
